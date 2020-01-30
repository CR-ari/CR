# -*- coding: utf-8 -*-
"""
Created on Wed Jan 22 16:23:34 2020

@author: 9876a
"""
import os
import sys
import openpyxl
import pandas as pd
import qrcode

path = input("Where is your file? Write your file's path")
list = os.listdir(path)
print(list)
#With this path, show the list of files in that path

fileName = input("Write your file name")
os.chdir(path)
excelFile = openpyxl.load_workbook(fileName)
#With file name, open the file

print(excelFile.sheetnames)
sheetName = input("Write the sheet name")
sheet = excelFile.get_sheet_by_name(sheetName)
#Decide which sheet you will open

df = pd.read_excel(fileName, sheet_name = sheetName)
row_number = len(df)
for_row = row_number + 2
column_number = (len(df.columns) + 1)
#Get the number of row and column
#Panda do not read the first row
#For using for loop, add 1 to each number of row and column

i = 1

for j in range(2,for_row):
    vaue = []
    for k in range(1,column_number):
        name = (str((sheet.cell(row=j, column=i)).value))
        vaue.append((str((sheet.cell(row=i, column=k).value)))+":"+(str((sheet.cell(row=j, column=k).value))))
    ima = qrcode.make(vaue)
    ima.save(name+'.png')    
#Make a list and put them on qrcode maker